import psycopg2
import openpyxl


class DatabaseManager:
    def __init__(self, dbname, user, password, host):
        self.dbname = dbname
        self.user = user
        self.password = password
        self.host = host

    def process_schedule(self, file_path):
        conn = psycopg2.connect(
            dbname=self.dbname,
            user=self.user,
            password=self.password,
            host=self.host
        )
        cur = conn.cursor()
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        for cell in sheet['B']:
            day_of_week = cell.value

            if day_of_week:
                cur.execute("SELECT EXISTS(SELECT 1 FROM weekday WHERE name = %s)", (day_of_week,))
                exists = cur.fetchone()[0]

                if not exists:
                    cur.execute("INSERT INTO weekday (name) VALUES (%s)", (day_of_week,))
                    conn.commit()

        for cell in sheet['C']:
            lesson_time = cell.value

            if lesson_time:
                cur.execute("SELECT EXISTS(SELECT 1 FROM lessons WHERE time = %s)", (lesson_time,))
                exists = cur.fetchone()[0]

                if not exists:
                    cur.execute("INSERT INTO lessons (time) VALUES (%s)", (lesson_time,))
                    conn.commit()

        for sheet_name in wb.sheetnames[:3]:
            sheet = wb[sheet_name]
            for row_num in range(10, 16):
                building_name = sheet[f'E{row_num}'].value
                building_color = sheet[f'E{row_num}'].fill.start_color.index if sheet[
                    f'E{row_num}'].fill.start_color else None
                if building_name:
                    cur.execute("SELECT EXISTS(SELECT 1 FROM address WHERE name = %s)", (building_name,))
                    exists = cur.fetchone()[0]

                    if not exists:
                        cur.execute("INSERT INTO address (name, color) VALUES (%s, %s)",
                                    (building_name, building_color))
                        conn.commit()

            for col in ['E', 'H', 'K']:
                group_name = sheet[f'{col}18'].value

                if group_name:
                    cur.execute("SELECT EXISTS(SELECT 1 FROM groups WHERE name = %s)", (group_name,))
                    exists = cur.fetchone()[0]

                    if not exists:
                        cur.execute("INSERT INTO groups (name) VALUES (%s)", (group_name,))
                        conn.commit()

            for col in ['F', 'I', 'L']:
                for row_num in range(20, sheet.max_row + 1):
                    classroom_name = sheet[f'{col}{row_num}'].value
                    classroom_color = sheet[f'{col}{row_num}'].fill.start_color.index if sheet[f'{col}{row_num}'].fill.start_color else None

                    if classroom_color == 'FFFFFFFF':
                        classroom_color = '00000000'

                    if classroom_name:
                        classroom_name = str(classroom_name).split(".")[0]

                        if classroom_color:
                            cur.execute("SELECT EXISTS(SELECT 1 FROM address WHERE color = %s)", (classroom_color,))
                            color_exists = cur.fetchone()[0]

                            if not color_exists:
                                continue

                        cur.execute("SELECT EXISTS(SELECT 1 FROM classroom WHERE name = %s AND color = %s)", (classroom_name, classroom_color))
                        exists = cur.fetchone()[0]

                        if not exists:
                            cur.execute("INSERT INTO classroom (name, color) VALUES (%s, %s)", (classroom_name, classroom_color))
                            conn.commit()

            for col in ['E', 'H', 'K']:
                for row_num in range(20, sheet.max_row + 1):
                    subject_name = sheet[f'{col}{row_num}'].value

                    if subject_name:
                        cur.execute("SELECT EXISTS(SELECT 1 FROM subject WHERE name = %s)", (subject_name,))
                        exists = cur.fetchone()[0]

                        if not exists:
                            cur.execute("INSERT INTO subject (name) VALUES (%s)", (subject_name,))
                            conn.commit()

            for col in ['E', 'H', 'K']:
                for row_num in range(20, sheet.max_row + 1):
                    subject_name = sheet[f'{col}{row_num}'].value

                    if subject_name:
                        cur.execute("SELECT id FROM subject WHERE name = %s", (subject_name,))
                        subject_id = cur.fetchone()[0]

                        classroom_name = sheet[f'{chr(ord(col) + 1)}{row_num}'].value
                        if classroom_name:
                            classroom_name = str(classroom_name).split(".")[0]
                            cur.execute("SELECT id FROM classroom WHERE name = %s", (classroom_name,))
                            classroom_id = cur.fetchone()[0]

                        lesson_time = sheet[f'C{row_num}'].value
                        if lesson_time:
                            cur.execute("SELECT id FROM lessons WHERE time = %s", (lesson_time,))
                            lesson_id = cur.fetchone()[0]

                        group_name = sheet[f'{col}18'].value
                        if group_name:
                            cur.execute("SELECT id FROM groups WHERE name = %s", (group_name,))
                            group_id = cur.fetchone()[0]

                        day_of_week = sheet[f'B{row_num}'].value
                        while not day_of_week and row_num > 1:
                            row_num -= 1
                            day_of_week = sheet[f'B{row_num}'].value

                        if day_of_week:
                            cur.execute("SELECT id FROM weekday WHERE name = %s", (day_of_week,))
                            weekday_id = cur.fetchone()[0]

                        cur.execute("""
                            INSERT INTO schedule (subject, classroom, lessons, groups, weekday, actuality, note)
                            SELECT %s, %s, %s, %s, %s, True, NULL
                            WHERE NOT EXISTS (
                                SELECT 1 FROM schedule
                                WHERE subject = %s AND classroom = %s AND lessons = %s AND groups = %s AND weekday = %s
                            )
                        """, (subject_id, classroom_id, lesson_id, group_id, weekday_id,
                              subject_id, classroom_id, lesson_id, group_id, weekday_id))

            cur.execute("""
                INSERT INTO main_schedule (id_sch)
                SELECT id FROM schedule
                WHERE actuality = %s AND NOT EXISTS (
                    SELECT 1 FROM main_schedule
                    WHERE id_sch = schedule.id
                )
            """, (True,))
            conn.commit()
            conn.commit()

        cur.close()
        conn.close()
        wb.close()